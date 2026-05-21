"""Excel (xlsx) 文档加载器（基于 openpyxl）

针对大表格优化：动态计算每组行数，确保不超过 embedding 模型的 token 限制。
每组带表头上下文，输出为 Markdown 表格格式。
"""

import os

from openpyxl import load_workbook

from app.pipeline.loader import BaseLoader, LoadResult

# embedding 模型最大输入字符数
_MAX_CHUNK_CHARS = 5000


class XlsxLoader(BaseLoader):
    """处理 .xlsx 文件的加载器

    策略：每个工作表按行分组，动态计算每组行数确保不超过模型限制。
    每组前面附带表头（第一行），输出为 Markdown 表格格式。
    """

    def load(self, file_path: str) -> LoadResult:
        """加载 Excel 文件，按行分组输出

        Args:
            file_path: 文件路径

        Returns:
            LoadResult: 包含文件内容和元数据
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        file_size = os.path.getsize(file_path)
        print(f"[XlsxLoader] 开始加载 Excel 文件: {os.path.basename(file_path)}, 大小: {file_size / 1024 / 1024:.1f}MB")

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"无法解析 xlsx 文件: {file_path}，错误: {e}")

        sheets_text = []
        total_rows = 0
        all_chunks = []

        for sheet in wb.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                rows.append(cells)

            if not rows:
                continue

            header = rows[0]
            data_rows = rows[1:]
            total_rows += len(data_rows)

            if not data_rows:
                sheet_text = f"[{sheet.title}]\n" + self._group_to_markdown(header, [])
                sheets_text.append(sheet_text)
                continue

            # 动态计算每组行数
            rows_per_chunk = self._calc_rows_per_chunk(header, data_rows)

            # 按行分组
            groups = []
            for i in range(0, len(data_rows), rows_per_chunk):
                group = data_rows[i:i + rows_per_chunk]
                groups.append(self._group_to_markdown(header, group))

            all_chunks.extend(groups)
            sheet_text = f"[{sheet.title}]\n\n" + "\n\n".join(groups)
            sheets_text.append(sheet_text)

        sheet_count = len(sheets_text)
        content = "\n\n".join(sheets_text)

        wb.close()

        print(f"[XlsxLoader] 加载完成，{sheet_count} 个工作表，共 {total_rows} 行数据，内容长度: {len(content)} 字符")

        metadata = {
            "filename": os.path.basename(file_path),
            "file_type": "xlsx",
            "file_size": file_size,
            "sheet_count": sheet_count,
            "row_count": total_rows,
        }

        return LoadResult(content=content, metadata=metadata, pre_chunked=all_chunks)

    def _calc_rows_per_chunk(self, header: list[str], data_rows: list[list[str]]) -> int:
        """根据实际数据动态计算每组行数"""
        clean_header = [cell.replace("\n", " ").replace("\r", " ").replace("|", "\\|") for cell in header]
        header_line = "| " + " | ".join(clean_header) + " |"
        sep_line = "| " + " | ".join(["---"] * len(header)) + " |"
        header_cost = len(header_line) + len(sep_line) + 2

        sample = data_rows[:200]
        if not sample:
            return 10

        row_widths = []
        for row in sample:
            col_count = len(header)
            padded = row + [""] * (col_count - len(row))
            padded = padded[:col_count]
            cells = [cell.replace("\n", " ").replace("\r", " ").replace("|", "\\|") for cell in padded]
            line = "| " + " | ".join(cells) + " |"
            row_widths.append(len(line) + 1)

        row_widths.sort()
        p90_idx = int(len(row_widths) * 0.9)
        p90_width = row_widths[p90_idx]

        available = _MAX_CHUNK_CHARS - header_cost
        if available <= 0:
            return 1

        rows_per_chunk = max(1, int(available / p90_width))
        rows_per_chunk = min(rows_per_chunk, 30)

        return rows_per_chunk

    def _group_to_markdown(self, header: list[str], rows: list[list[str]]) -> str:
        """将一组行（带表头）转为 Markdown 表格"""
        col_count = len(header)
        lines = []
        clean_header = [cell.replace("\n", " ").replace("\r", " ").replace("|", "\\|") for cell in header]
        lines.append("| " + " | ".join(clean_header) + " |")
        lines.append("| " + " | ".join(["---"] * col_count) + " |")

        for row in rows:
            padded = row + [""] * (col_count - len(row))
            padded = padded[:col_count]
            cells = [cell.replace("\n", " ").replace("\r", " ").replace("|", "\\|") for cell in padded]
            lines.append("| " + " | ".join(cells) + " |")

        return "\n".join(lines)
